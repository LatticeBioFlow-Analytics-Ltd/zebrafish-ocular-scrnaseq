"""Article figure panels: Fig. 2C, 2D and S3.

Panels for the transporter manuscript, built on the pass-2 ocular subset:

  2C  UMAP with the RPE cluster identified by rpe65a/tyrp1b/pmela, both stages.
  2D  Control-panel dotplot: the three target transporters bracketed by
      positive and negative controls, across every ocular cell type at both
      stages. Dot size is the DETECTION FRACTION and colour the mean
      expression among expressing cells only - for low-abundance transcripts
      the detection fraction is the honest statistic, and a mean taken across
      all cells conflates abundance with dropout.
  S3  Supplementary: cluster marker dotplot, cells per cluster per stage,
      and RPE vs non-RPE detection fractions for the whole panel, plus an
      auto-evaluated acceptance-criteria table.

The gene panel, its tiered roles and the acceptance criteria are transcribed
from `transporter_dotplot_control_panel.xlsx` into `config/article_panel.yaml`;
the workbook is the authority on why each gene is on the plot. Two of its
criteria cannot be evaluated from this repository alone and are recorded as
pending in the S3 notes: ambient-RNA correction (run the pipeline on the
CellBender tier via `config/config_snakemake.yaml` once the upstream Snakemake
workflow has produced it) and replication in a second public atlas.
"""

from __future__ import annotations

import argparse
import logging
import sys
import textwrap
from pathlib import Path

import anndata as ad
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scanpy as sc
import yaml

from . import figures as figs
from .pipeline import (
    DOT_EDGE_LW,
    LARGEST_DOT,
    PANEL_SPACING,
    SMALLEST_DOT,
    _marker_dotplot,
    headless_plotting,
)

log = logging.getLogger(__name__)


def load_panel_spec(path: Path) -> dict:
    """Read config/article_panel.yaml."""
    with Path(path).open() as handle:
        return yaml.safe_load(handle)


def _rpe_mask(adata: ad.AnnData, label: str) -> np.ndarray:
    """Cells labelled as the RPE cluster, including the low-confidence variant."""
    ct = adata.obs["cell_type"].astype(str)
    return ((ct == label) | (ct == f"{label} (low confidence)")).to_numpy()


def _present(adata: ad.AnnData, genes: list[str]) -> tuple[list[str], list[str]]:
    """Split genes into (present, missing) against the full .raw gene space."""
    var = set(adata.raw.var_names)
    present = [g for g in genes if g in var]
    missing = [g for g in genes if g not in var]
    if missing:
        log.warning(
            "Article panel genes not in the reference: %s. Check the symbol "
            "against this annotation release - a control absent for naming "
            "reasons is worse than no control at all.", ", ".join(missing),
        )
    return present, missing


def _detected(adata: ad.AnnData, gene: str) -> np.ndarray:
    """Boolean per cell: any transcript of `gene` detected."""
    column = adata.raw[:, gene].X
    dense = column.toarray().ravel() if hasattr(column, "toarray") else np.asarray(column).ravel()
    return dense > 0


def detection_table(adata: ad.AnnData, spec: dict) -> pd.DataFrame:
    """Detection fraction per panel gene, RPE vs non-RPE, per stage and pooled.

    These are the numbers the acceptance criteria compare (and the workbook's
    fill-in columns H-K): the fraction of cells with any detected transcript,
    not mean expression.
    """
    rpe = _rpe_mask(adata, spec["rpe_cell_type"])
    stages = list(adata.obs["timepoint"].cat.categories)
    rows = []
    for role, genes in spec["panel"].items():
        _, missing = _present(adata, genes)
        for gene in genes:
            if gene in missing:
                rows.append({"gene": gene, "role": role, "stage": "pooled",
                             "detected_in_dataset": False, "pct_rpe": np.nan,
                             "pct_non_rpe": np.nan, "n_rpe": int(rpe.sum()),
                             "n_non_rpe": int((~rpe).sum())})
                continue
            hit = _detected(adata, gene)
            for stage in [*stages, "pooled"]:
                in_stage = (np.ones(adata.n_obs, bool) if stage == "pooled"
                            else (adata.obs["timepoint"] == stage).to_numpy())
                r, o = rpe & in_stage, ~rpe & in_stage
                rows.append({
                    "gene": gene, "role": role, "stage": stage,
                    "detected_in_dataset": bool(hit.any()),
                    "pct_rpe": 100 * hit[r].mean() if r.any() else np.nan,
                    "pct_non_rpe": 100 * hit[o].mean() if o.any() else np.nan,
                    "n_rpe": int(r.sum()), "n_non_rpe": int(o.sum()),
                })
    return pd.DataFrame(rows)


# --- Fig 2C -------------------------------------------------------------------

def fig2c(adata: ad.AnnData, spec: dict, fig_dir: Path) -> None:
    """UMAP with the RPE cluster identified, per stage, plus identity markers."""
    label = spec["rpe_cell_type"]
    groups = [c for c in adata.obs["cell_type"].cat.categories
              if str(c) == label or str(c) == f"{label} (low confidence)"]
    if not groups:
        log.warning("fig2C skipped: no %r cell type in this subset", label)
        return

    stages = list(adata.obs["timepoint"].cat.categories)
    # One point size for every panel, derived from the full subset: scanpy's
    # per-plot default (120000/n) would render the sparse 8 dpf panel as
    # boulders next to the 5 dpf gravel.
    size = 120000 / adata.n_obs
    _fig, axes = plt.subplots(1, len(stages),
                              figsize=(figs.DOUBLE_COL, figs.DOUBLE_COL / len(stages) * 0.95))
    for ax, stage in zip(np.atleast_1d(axes), stages):
        sub = adata[(adata.obs["timepoint"] == stage).to_numpy()].copy()
        # The RPE cluster gets an explicit colour: its registered palette
        # entry can be Okabe-Ito black, which reads as "greyed out" next to
        # the light-grey background cells rather than as the highlight.
        cats = list(sub.obs["cell_type"].cat.categories)
        colours = list(sub.uns.get("cell_type_colors",
                                   figs.CATEGORICAL[:len(cats)]))
        sub.uns["cell_type_colors"] = [
            "#B2182B" if c in groups else col for c, col in zip(cats, colours)
        ]
        n = int(_rpe_mask(sub, label).sum())
        # `groups` greys every other cluster; the count in the title is
        # deliberate - the acceptance criteria require the RPE cluster size
        # per stage to be stated, since a few dozen cells cannot support
        # claims about low-abundance transcripts.
        sc.pl.umap(sub, color="cell_type", groups=groups, ax=ax, show=False,
                   frameon=False, legend_loc="none", size=size,
                   title=f"{stage}  ({n} RPE cells)", na_in_legend=False)
    figs.save_figure(fig_dir, "fig2C_rpe_by_stage")

    present, _ = _present(adata, spec["rpe_identity_markers"])
    if present:
        figs.scanpy_panel_size(n_cols=3)
        sc.pl.umap(adata, color=present, use_raw=True, ncols=3,
                   cmap=figs.SEQUENTIAL, vmin=0, vmax="p99", frameon=False,
                   wspace=PANEL_SPACING, show=False)
        figs.save_figure(fig_dir, "fig2C_rpe_markers")

    (fig_dir / "fig2C_legend.txt").write_text(
        "Fig 2C. UMAP of the ocular subset with the RPE cluster coloured and "
        "all other cells grey, shown separately for each stage; the RPE cell "
        "count per stage is given in each panel title. The companion panel "
        f"shows the identity markers ({', '.join(present)}) that define the "
        "cluster, coloured by log-normalised expression (99th-percentile "
        "ceiling).\n"
    )


# --- Fig 2D -------------------------------------------------------------------

def fig2d(adata: ad.AnnData, spec: dict, fig_dir: Path) -> None:
    """Control-panel dotplot across cell types and stages.

    Dot size encodes the detection fraction on an ABSOLUTE 0-100% scale
    (`dot_max=1`), so a target detected in 5% of RPE cells is visibly smaller
    than an ambient sentinel at 90% - rescaling to the column maximum would
    manufacture parity. Colour is the mean among expressing cells only
    (`mean_only_expressed`), which separates "how strongly, where seen" from
    "how often seen" instead of blending them.
    """
    display: dict[str, list[str]] = {}
    for role, genes in spec["panel"].items():
        present, _ = _present(adata, genes)
        if present:
            display[textwrap.fill(role, width=14)] = present
    if not display:
        log.warning("fig2D skipped: no panel gene present")
        return

    key = "_article_row"
    stages = list(adata.obs["timepoint"].cat.categories)
    order = [f"{ct} · {tp}" for ct in adata.obs["cell_type"].cat.categories
             for tp in stages]
    rows = (adata.obs["cell_type"].astype(str) + " · "
            + adata.obs["timepoint"].astype(str))
    adata.obs[key] = pd.Categorical(rows, categories=[o for o in order
                                                      if o in set(rows)])
    try:
        plot = sc.pl.dotplot(
            adata, display, groupby=key, use_raw=True,
            mean_only_expressed=True, dot_max=1.0,
            cmap=figs.SEQUENTIAL, return_fig=True, show=False,
        )
        plot.DEFAULT_CATEGORY_WIDTH = figs.DOT_PITCH
        plot.DEFAULT_CATEGORY_HEIGHT = figs.DOT_PITCH
        plot.style(
            largest_dot=LARGEST_DOT, smallest_dot=SMALLEST_DOT,
            dot_edge_lw=DOT_EDGE_LW, x_padding=0.45, y_padding=0.45,
        ).legend(width=figs.DOT_LEGEND_WIDTH,
                 colorbar_title="Mean expression\nin expressing cells",
                 size_title="Cells detecting\nthe gene (%)")
        plot.make_figure()
        # Gene symbols are italicised by convention; scanpy has no hook for it.
        main = plot.get_axes().get("mainplot_ax")
        if main is not None:
            for tick in main.get_xticklabels():
                tick.set_fontstyle("italic")
        figs.save_figure(fig_dir, "fig2D_transporter_controls", width=None)
    finally:
        del adata.obs[key]

    (fig_dir / "fig2D_legend.txt").write_text(
        "Fig 2D. Detection of the target transporters (slc39a14, slc39a8, "
        "slc30a10) across every ocular cell type at 5 and 8 dpf, bracketed by "
        "controls chosen to span a range of expression levels: RPE identity "
        "markers, abundance-matched positives (slc11a2, best1), ambient-RNA "
        "sentinels (rho, gnat2, elavl3) and out-of-tissue negatives (cryaa, "
        "hbae1.1). Dot size shows the fraction of cells in which the gene is "
        "detected, on an absolute scale to 100%; colour shows mean "
        "log-normalised expression among expressing cells only. Detection "
        "fraction rather than mean expression is shown because, for "
        "low-abundance transcripts, a mean across all cells conflates "
        "abundance with dropout.\n"
    )


# --- Fig S3 -------------------------------------------------------------------

def _plot_cells_per_cluster(adata: ad.AnnData, fig_dir: Path) -> None:
    counts = pd.crosstab(adata.obs["cell_type"], adata.obs["timepoint"])
    stages = list(counts.columns)
    _fig, ax = figs.figure(width=figs.SINGLE_COL,
                           height=0.6 + 0.34 * len(counts))
    y = np.arange(len(counts))
    height = 0.8 / len(stages)
    for j, stage in enumerate(stages):
        offset = (j - (len(stages) - 1) / 2) * height
        ax.barh(y + offset, counts[stage], height=height * 0.9,
                color=figs.CATEGORICAL[j % len(figs.CATEGORICAL)], label=stage)
        for k, v in enumerate(counts[stage]):
            ax.text(max(v, 1) * 1.15, k + offset, f"{v:,}",
                    va="center", fontsize=6)
    ax.set_xscale("log")   # 51k- and 8-cell bars must both be readable
    ax.set_xlim(left=1)
    ax.set_yticks(y)
    ax.set_yticklabels([textwrap.fill(str(t).replace("_", " "), 20)
                        for t in counts.index])
    ax.invert_yaxis()
    ax.set_xlabel("Cells (log scale)")
    ax.set_title("Ocular cells per cluster and stage", loc="left")
    # Below the x-label, as in the composition plot: bars can reach any height
    # on the right, so an in-axes legend has nowhere reliable to sit.
    axes_height = _fig.get_size_inches()[1] * ax.get_position().height
    ax.legend(frameon=False, loc="upper center", ncols=len(stages),
              bbox_to_anchor=(0.5, -0.45 / axes_height), borderaxespad=0)
    figs.save_figure(fig_dir, "figS3_cells_per_cluster", width=figs.SINGLE_COL)


def _plot_detection_fractions(table: pd.DataFrame, spec: dict,
                              fig_dir: Path) -> None:
    genes = [g for gs in spec["panel"].values() for g in gs]
    stages = [s for s in table["stage"].unique() if s != "pooled"]
    _fig, axes = plt.subplots(1, len(stages),
                              figsize=(figs.DOUBLE_COL, 0.5 + 0.17 * len(genes)),
                              sharey=True)
    y = np.arange(len(genes))
    for ax, stage in zip(np.atleast_1d(axes), stages):
        sub = table[table["stage"] == stage].set_index("gene")
        rpe = [sub["pct_rpe"].get(g, np.nan) for g in genes]
        oth = [sub["pct_non_rpe"].get(g, np.nan) for g in genes]
        n = int(sub["n_rpe"].max()) if len(sub) else 0
        for k in y:
            ax.hlines(k, 0, np.nanmax([rpe[k], oth[k], 0]), color="0.85",
                      linewidth=0.5, zorder=1)
        ax.scatter(oth, y, s=14, facecolor="white", edgecolor="0.35",
                   linewidth=0.6, zorder=3, label="all other ocular cells")
        ax.scatter(rpe, y, s=16, color="#B2182B", edgecolor="black",
                   linewidth=0.3, zorder=4, label="RPE cells")
        ax.set_yticks(y)
        ax.set_yticklabels(genes, fontstyle="italic")
        ax.set_xlim(left=0)
        ax.set_xlabel("Cells detecting the gene (%)")
        ax.set_title(f"{stage}  ({n} RPE cells)", loc="left")
    axes = np.atleast_1d(axes)
    # Inverted once, outside the loop: with sharey=True a per-panel
    # invert_yaxis() toggles the shared axis and an even panel count lands
    # back in bottom-up order.
    axes[0].invert_yaxis()
    axes[-1].legend(frameon=False, loc="lower right", handletextpad=0.3)
    figs.save_figure(fig_dir, "figS3_detection_fractions")


def _acceptance_criteria(table: pd.DataFrame, spec: dict) -> pd.DataFrame:
    """Auto-evaluate the workbook's objective acceptance criteria.

    Only the criteria computable from this dataset get a verdict; "comparable"
    judgements are reported as values for a human to weigh, and the two
    criteria needing external inputs (ambient correction, second atlas) are
    marked pending.
    """
    pooled = table[table["stage"] == "pooled"].set_index("gene")
    per_stage = {s: table[table["stage"] == s].set_index("gene")
                 for s in table["stage"].unique() if s != "pooled"}
    targets = spec["panel"]["target"]
    sentinel = "rho"
    matched = spec["panel"]["abundance-matched positive"][0]
    out_of_tissue = spec["panel"]["out-of-tissue"]

    def pct(frame, gene, col="pct_rpe"):
        return float(frame[col].get(gene, np.nan))

    rows = []
    exceeds = {t: pct(pooled, t) > pct(pooled, sentinel) for t in targets}
    rows.append({
        "criterion": f"1. Each target detected in more RPE cells than {sentinel}",
        "value": "; ".join(f"{t} {pct(pooled, t):.1f}% vs {sentinel} "
                           f"{pct(pooled, sentinel):.1f}%" for t in targets),
        "verdict": "pass" if all(exceeds.values()) else "FAIL",
    })
    rows.append({
        "criterion": f"2. Target detection comparable to {matched}",
        "value": "; ".join(f"{t} {pct(pooled, t):.1f}% vs {matched} "
                           f"{pct(pooled, matched):.1f}%" for t in targets),
        "verdict": "review",   # "comparable" is a judgement, not a threshold
    })
    enriched = {t: pct(pooled, t) > pct(pooled, t, "pct_non_rpe") for t in targets}
    rows.append({
        "criterion": "3. Detection enriched in RPE vs other clusters",
        "value": "; ".join(f"{t} RPE {pct(pooled, t):.1f}% vs other "
                           f"{pct(pooled, t, 'pct_non_rpe'):.1f}%" for t in targets),
        "verdict": "pass" if all(enriched.values()) else "FAIL",
    })
    rows.append({
        "criterion": "4. Ambient-RNA correction applied and reported",
        "value": "not applied in this run; see figS3_notes.txt",
        "verdict": "pending",
    })
    sizes = {s: int(f["n_rpe"].max()) for s, f in per_stage.items()}
    rows.append({
        "criterion": "5. RPE cluster size stated per stage",
        "value": "; ".join(f"{s}: {n} cells" for s, n in sizes.items()),
        "verdict": "pass" if min(sizes.values()) >= 50
                   else "FAIL (a cluster this small cannot support claims "
                        "about low-abundance transcripts)",
    })
    both = all(
        pct(f, t) > pct(f, sentinel) and pct(f, t) > pct(f, t, "pct_non_rpe")
        for f in per_stage.values() for t in targets
    )
    rows.append({
        "criterion": "6. Result reproduces at both stages independently",
        "value": "; ".join(
            f"{s}: " + ", ".join(f"{t} {pct(f, t):.1f}%" for t in targets)
            for s, f in per_stage.items()),
        # Replication needs two stages to compare; a single-stage run cannot
        # fail this criterion, it just cannot address it.
        "verdict": ("pending (single stage in this run)" if len(per_stage) < 2
                    else "pass" if both else "FAIL"),
    })
    rows.append({
        "criterion": "7. Result reproduces in a second public atlas",
        "value": "not evaluated here; see figS3_notes.txt",
        "verdict": "pending",
    })
    worst = {}
    for g in out_of_tissue:
        vals = [v for v in (pct(pooled, g), pct(pooled, g, "pct_non_rpe"))
                if not np.isnan(v)]
        worst[g] = max(vals) if vals else float("nan")  # nanmax warns on all-NaN
    rows.append({
        "criterion": "8. Out-of-tissue controls near zero in every cluster",
        "value": "; ".join(f"{g} max {v:.1f}%" for g, v in worst.items()),
        "verdict": "pass" if all(v < 5 for v in worst.values()) else "review",
    })
    return pd.DataFrame(rows)


def figs3(adata: ad.AnnData, spec: dict, markers: dict[str, list[str]],
          fig_dir: Path, table_dir: Path) -> pd.DataFrame:
    """Supplementary figure S3 and its tables. Returns the detection table."""
    present_markers = {
        name: [g for g in genes if g in adata.raw.var_names]
        for name, genes in markers.items()
    }
    present_markers = {k: v for k, v in present_markers.items() if v}
    if present_markers:
        _marker_dotplot(adata, present_markers, "cell_type", fig_dir,
                        "figS3_marker_dotplot")

    _plot_cells_per_cluster(adata, fig_dir)

    table = detection_table(adata, spec)
    table.to_csv(table_dir / "figS3_detection_fractions.csv", index=False)
    _plot_detection_fractions(table, spec, fig_dir)

    criteria = _acceptance_criteria(table, spec)
    criteria.to_csv(table_dir / "figS3_acceptance_criteria.csv", index=False)
    n_fail = int(criteria["verdict"].str.startswith("FAIL").sum())
    if n_fail:
        log.warning("%d acceptance criterion(s) FAIL - see "
                    "figS3_acceptance_criteria.csv", n_fail)

    (fig_dir / "figS3_notes.txt").write_text(
        "Fig S3 notes.\n\n"
        "Ambient-RNA correction (acceptance criterion 4): NOT applied in this "
        "run. The inputs are the uncorrected per-timepoint aggregates. The "
        "upstream scRNA_analysis Snakemake workflow publishes per-library "
        "CellBender-corrected matrices; re-running this pipeline with "
        "config/config_snakemake.yaml uses those as input, after which every "
        "panel here is ambient-corrected. Until then, treat the ambient "
        "sentinels (rho, gnat2, elavl3) as the contamination readout.\n\n"
        "Second public atlas (criterion 7): not evaluated in this repository. "
        "Candidates named by the control workbook: Daniocell (Sur et al. "
        "2023) and Farnsworth et al. 2020; check their RPE clusters for the "
        "three transporters.\n\n"
        "Gene naming: hbae1 does not exist in this reference (Ensembl 116 "
        "GRCz11); hbae1.1 is plotted in its place as the out-of-tissue "
        "erythrocyte control, per the workbook's warning that paralogue "
        "naming must be verified against the annotation release.\n\n"
        "Detection fractions are computed on the log-normalised .raw matrix "
        "(a cell counts as detecting a gene if it has any non-zero count).\n"
    )
    return table


def fill_workbook(table: pd.DataFrame, criteria: pd.DataFrame,
                  src: Path, dst: Path) -> None:
    """Fill the control workbook's yellow columns (H-K) and criteria sheet.

    Writes a COPY next to the other article outputs; the user's original stays
    untouched. Symbols are matched via the same aliasing the panel config uses
    (the sheet says `hbae1`, this annotation has `hbae1.1`), and the
    substitution is recorded in the row's Notes cell.
    """
    from openpyxl import load_workbook

    pooled = table[table["stage"] == "pooled"].set_index("gene")
    stages = [s for s in table["stage"].unique() if s != "pooled"]
    per_stage = {s: table[table["stage"] == s].set_index("gene") for s in stages}
    aliases = {"hbae1": "hbae1.1"}

    wb = load_workbook(src)
    ws = wb["Control panel"]
    for row in ws.iter_rows(min_row=2):
        symbol = row[1].value  # column B
        gene = aliases.get(symbol, symbol)
        if not symbol or gene not in pooled.index:
            continue
        entry = pooled.loc[gene]
        row[7].value = "Y" if bool(entry["detected_in_dataset"]) else "N"  # H
        row[8].value = round(float(entry["pct_rpe"]), 1)                   # I
        row[9].value = round(float(entry["pct_non_rpe"]), 1)               # J
        note = "; ".join(
            f"{s}: RPE {per_stage[s]['pct_rpe'].get(gene, float('nan')):.1f}%"
            f" (n={int(per_stage[s]['n_rpe'].max())}), other "
            f"{per_stage[s]['pct_non_rpe'].get(gene, float('nan')):.1f}%"
            for s in stages
        )
        if symbol in aliases:
            note = f"plotted as {gene} (symbol absent in this annotation); " + note
        row[10].value = note                                               # K

    ws = wb["Acceptance criteria"]
    verdicts = {int(c.split(".")[0]): (c, v, val) for c, v, val in zip(
        criteria["criterion"], criteria["verdict"], criteria["value"])}
    for row in ws.iter_rows(min_row=2):
        try:
            number = int(row[0].value)
        except (TypeError, ValueError):
            continue
        if number in verdicts:
            _, verdict, value = verdicts[number]
            row[3].value = verdict   # D: Pass / fail
            row[4].value = value     # E: Value
    wb.save(dst)
    log.info("Filled control workbook written to %s", dst)


def plot_article_panels(adata: ad.AnnData, spec_path: Path,
                        markers_path: Path, out_dir: Path) -> None:
    """All article panels into `<out_dir>/figures` + tables into `<out_dir>`."""
    from .cluster import load_markers

    spec = load_panel_spec(spec_path)
    fig_dir = out_dir / "figures"
    fig_dir.mkdir(parents=True, exist_ok=True)

    fig2c(adata, spec, fig_dir)
    fig2d(adata, spec, fig_dir)
    figs3(adata, spec, load_markers(markers_path), fig_dir, out_dir)
    log.info("Article panels written to %s", out_dir)


def main(argv: list[str] | None = None) -> int:
    """Standalone entry point: build the panels from a saved ocular subset.

    Reads the pass-2 output rather than re-running the pipeline, so panel
    iteration costs seconds, not the ~35 minutes of a full run.
    """
    from .cli import find_repo_root

    parser = argparse.ArgumentParser(
        description="Build article figure panels (2C, 2D, S3) from the "
                    "pass-2 ocular subset.")
    parser.add_argument("--ocular", default="results/ocular/ocular.h5ad")
    parser.add_argument("--panel", default="config/article_panel.yaml")
    parser.add_argument("--markers", default="config/markers_ocular.yaml")
    parser.add_argument("--outdir", default="results/article")
    parser.add_argument("--stage", default=None,
                        help="restrict every panel to one timepoint (e.g. "
                             "5dpf); pair it with a distinct --outdir so the "
                             "single-stage set never overwrites the full one")
    parser.add_argument("--xlsx", default=None,
                        help="control workbook to fill in (columns H-K and "
                             "the acceptance-criteria sheet); a filled copy "
                             "is written to the output directory")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO,
                        format="%(levelname)-7s | %(message)s")
    root = find_repo_root()
    figs.apply_style()
    sc.settings.autoshow = False

    adata = ad.read_h5ad(root / args.ocular)
    if args.stage:
        keep = (adata.obs["timepoint"] == args.stage).to_numpy()
        if not keep.any():
            available = list(adata.obs["timepoint"].cat.categories)
            log.error("No cells at stage %r; available: %s", args.stage, available)
            return 1
        adata = adata[keep].copy()
        # Drop the empty categories, or every per-stage loop still draws an
        # empty panel (and the acceptance criteria a 0% stage) for the
        # timepoint that was just filtered out.
        adata.obs["timepoint"] = adata.obs["timepoint"].cat.remove_unused_categories()
        adata.obs["sample"] = adata.obs["sample"].cat.remove_unused_categories()
        log.info("Restricted to %s: %d cells", args.stage, adata.n_obs)

    with headless_plotting():
        plot_article_panels(adata, root / args.panel, root / args.markers,
                            root / args.outdir)

    if args.xlsx:
        spec = load_panel_spec(root / args.panel)
        table = detection_table(adata, spec)
        criteria = _acceptance_criteria(table, spec)
        src = Path(args.xlsx).expanduser()
        fill_workbook(table, criteria, src,
                      root / args.outdir / f"{src.stem}_filled.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())
